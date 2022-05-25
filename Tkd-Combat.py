from tkinter import *
from tkinter import ttk, filedialog
from tkinter.messagebox import *
from typing import Type
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import xlsxwriter as xw 
import pandas as pd


class App(Tk):

    def __init__(self):
        Tk.__init__(self)
        self.geometry("1000x450")
        self.title('Tkd Competitors')
        self.iconbitmap('tkd.ico')
        self.resizable(0,0)

        self._frame = None
        self.switch_frame(Competidor)

    def switch_frame(self, frame_class):
        new_frame = frame_class(self)
        if self._frame is not None:
            self._frame.destroy()
        self._frame = new_frame
        self._frame.pack()
        
# Agregar Competidor
class Competidor(Frame):
    def __init__(self, master):

        Frame.__init__(self,master)

        # Frame buttons left
        btn_frame = Frame(self, bg='#868B8E', width=200, height=800)
        btn_frame.pack(padx=0, pady=0, side='left')
        btn_frame.grid_propagate(0)

        # Frame Labels and buttons
        self.label_frame = Frame(self, bg='#145DA0', width=800, height=800)
        self.label_frame.pack(padx=0, pady=0, side='left')
        self.label_frame.grid_propagate(0)   

        #Buttons Opntions
        btn_agr = Button(btn_frame, text='Agregar Competidor/a', bd=5, command=lambda: master.switch_frame(Competidor))
        btn_agr.grid(row=0, column=0, padx=25, pady=80, ipadx=5, ipady= 5)

        btn_list = Button(btn_frame, text='Listar Competidores/as', bd=5, command=lambda: master.switch_frame(ListCompet))
        btn_list.grid(row=1, column=0, padx=25, pady=(0,40), ipadx=5, ipady= 5)

        btn_word = Button(btn_frame, text='Crear Llaves', bd=5, command=lambda: master.switch_frame(CreateKeys))
        btn_word.grid(row=2, column=0, padx=25, pady=30, ipadx=5, ipady= 5)

        self._createWidget()

 
    def _createWidget(self):

        self.optSex = StringVar()

        def only_numbers(char):
            return char.isdigit()

        validation = self.label_frame.register(only_numbers)

        #Nombre
        nmc= Label(self.label_frame,text="Nombre Competidor/a: ", bg='#145DA0', font=(10))
        nmc.grid(row=0, column=0, pady=50, padx=(50,0))

        self.nmcE = Entry(self.label_frame,font=(10))
        self.nmcE.grid( row=0, column=1, padx=20)

        #Sexo
        sxc = ttk.LabelFrame(self.label_frame,text="Sexo")
        sxc.grid(row=0, column=2, pady=(60,0) ,padx=(100,0))

        self.rbSex = ttk.Radiobutton(sxc, variable=self.optSex, value='M' ,text='Masculino')
        self.rbSex.grid(row=0, column=3, padx=20, pady=10)

        self.rbSex2 = ttk.Radiobutton(sxc, variable=self.optSex, value='F',text='Femenino')
        self.rbSex2.grid(row=1, column=3, padx=20, pady=10)

        #Edad
        edc = Label(self.label_frame, text="Edad Competidor/a: ", bg='#145DA0', font=(10))
        edc.grid(row=1, column=0, pady=0, padx=(50,0))

        self.edadE = Entry(self.label_frame,font=(10), width=8, validate='key', validatecommand=(validation, '%S'))
        self.edadE.grid( row=1, column=1, padx=20)
        
        #Peso
        psc = Label(self.label_frame, text="Peso Competidor/a:", bg='#145DA0', font=(10))
        psc.grid(row=2, column=0, pady=60, padx=25)

        self.pscE = Entry(self.label_frame, font=(10), width=8)
        self.pscE.grid(row=2, column=1, padx=20)
        
        #Color Cinturon
        cinturones = ['Blanco','Blanco-Amarillo','Amarillo','Amarillo-Verde',
                     'Verde','Verde-Azul','Azul','Azul-Rojo','Rojo','Rojo-Negro',
                     'Negro']
        
        clc = Label(self.label_frame,text="Color Cinturon: ", bg='#145DA0', font=(10))
        clc.grid(row=3,column=0,pady=(0,10))

        self.clcbx = ttk.Combobox(self.label_frame,values=cinturones, state='readonly',font=(10), width=14)
        self.clcbx.grid(row=3, column=1, padx=20)
        self.clcbx.current(0)

        #Button
        btn_sv = Button(self.label_frame,text="Guardar", bd=5, command=self.save_data)
        btn_sv.grid(row=4,column=2, padx=20,ipady=5,ipadx=20)

    # Save_Data es la funcion que va a guardar los datos en excel
    def save_data(self):
        
        clc = self.clcbx.get()
        
        try:
            nm = str(self.nmcE.get())
            if nm.isdigit() :
                raise ValueError()
            elif len(nm) == 0:
                raise ValueError()
        except ValueError:
            (showerror("Error","Debe Escribir el Nombre del/a competidor/a Correctamente"))
        try: 
            sex = self.optSex.get()
            if sex == '':
                raise ValueError()
        except ValueError:
            showerror("Error","Debe Seleccionar el sexo del/a competidor/a")
        try:
            edad = int(self.edadE.get())
        except ValueError:
            showerror("Error","Debe Escribir la edad del/a competidor/a")
        try:
            peso = float(self.pscE.get())
        except ValueError:
            showerror("Error","Debe Escribir el peso del/a competidor/a")
        try:
            wb = load_workbook("Listado_de_Competidores_Kyrugi.xlsx")
            datos = (
                ['Nombre','Edad','Peso','Sexo','Color Cinturon'],
                [nm, edad, peso, sex, clc]
            )
            print(datos)

            ws = wb["Listado General de Competidores"]
            wb.active = ws
            current_row = ws.max_row

            ws.cell(row=current_row + 1, column=1).value = nm
            ws.cell(row=current_row + 1, column=2).value = edad
            ws.cell(row=current_row + 1, column=3).value = sex
            ws.cell(row=current_row + 1, column=4).value = peso
            ws.cell(row=current_row + 1, column=5).value = clc

            wb.save('Listado_de_Competidores_Kyrugi.xlsx')  
            print(f'Nombre de hoja: {wb.active.title}')


            if peso <= 79:
                ws1 = wb["Categoria Menor a 80Kg"]
                wb.active = ws1
                current_row = ws1.max_row

                ws1.cell(row=current_row + 1, column=1).value = nm
                ws1.cell(row=current_row + 1, column=2).value = edad
                ws1.cell(row=current_row + 1, column=3).value = sex
                ws1.cell(row=current_row + 1, column=4).value = peso
                ws1.cell(row=current_row + 1, column=5).value = clc

                wb.save('Listado_de_Competidores_Kyrugi.xlsx')  
                print(f'Nombre de hoja: {wb.active.title}')
                
            if peso >= 80:
                ws2 = wb["Categoria Mayor a 80Kg"]
                wb.active = ws2
                current_row = ws2.max_row

                ws2.cell(row=current_row + 1, column=1).value = nm
                ws2.cell(row=current_row + 1, column=2).value = edad
                ws2.cell(row=current_row + 1, column=3).value = sex
                ws2.cell(row=current_row + 1, column=4).value = peso
                ws2.cell(row=current_row + 1, column=5).value = clc

                wb.save('Listado_de_Competidores_Kyrugi.xlsx')  
                print(f'Nombre de hoja: {wb.active.title}')
            

            self.nmcE.delete(0,END),
            self.edadE.delete(0,END),
            self.pscE.delete(0,END),
            self.optSex.set(None),
            self.clcbx.current(0)

        except IOError:
            showinfo('Informacion', 'Se esta creando el archivo excel. Guarde nuevamente al competidor.')
            wb = Workbook()
            ws = wb.active

            ws.title = "Listado General de Competidores"
            ws['A1'] = "NOMBRE COMPLETO"
            ws['B1'] = "EDAD"
            ws['C1'] = "SEXO"
            ws['D1'] = "PESO"
            ws['E1'] = "COLOR CINTURON"
            

            ws1 = wb.create_sheet("Categoria Menor a 80Kg")

            ws1['A1'] = "NOMBRE COMPLETO"
            ws1['B1'] = "EDAD"
            ws1['C1'] = "SEXO"
            ws1['D1'] = "PESO"
            ws1['E1'] = "COLOR CINTURON"

            ws2 = wb.create_sheet("Categoria Mayor a 80Kg")

            ws2['A1'] = "NOMBRE COMPLETO"
            ws2['B1'] = "EDAD"
            ws2['C1'] = "SEXO"
            ws2['D1'] = "PESO"
            ws2['E1'] = "COLOR CINTURON"



            wb.save('Listado_de_Competidores_Kyrugi.xlsx')


class ListCompet(Frame):
    def __init__(self,master):
        Frame.__init__(self,master)

        # Frame buttons left
        btn_frame = Frame(self, bg='#868B8E', width=200, height=800)
        btn_frame.pack(padx=0, pady=0, side='left')
        btn_frame.grid_propagate(0)

        # Frame Labels and buttons
        self.label_frame = Frame(self, bg='#A0E7E5', width=800, height=800)
        self.label_frame.pack(padx=0, pady=0, side='left')
        self.label_frame.grid_propagate(0)   

        #Buttons Opntions
        btn_agr = Button(btn_frame, text='Agregar Competidor/a', bd=5, command=lambda: master.switch_frame(Competidor))
        btn_agr.grid(row=0, column=0, padx=25, pady=80, ipadx=5, ipady= 5)

        btn_list = Button(btn_frame, text='Listar Competidores/as', bd=5, command=lambda: master.switch_frame(ListCompet))
        btn_list.grid(row=1, column=0, padx=25, pady=(0,40), ipadx=5, ipady= 5)

        btn_word = Button(btn_frame, text='Crear Llaves', bd=5, command=lambda: master.switch_frame(CreateKeys))
        btn_word.grid(row=2, column=0, padx=25, pady=30, ipadx=5, ipady= 5)

        self.WidgetCompet()

    def WidgetCompet(self):
        
        frame1 = LabelFrame(self.label_frame, text='Informacion Excel')
        frame1.place(height=300, width=800, rely=0.04)

        listado_frame = LabelFrame(self.label_frame, text='listado de Hojas')
        listado_frame.place(height=100, width=200, rely=0.75, relx=0.65)

        sheet = ttk.Combobox(listado_frame, state='readonly')
        sheet.grid(row=1, column=1, padx=20)
        sheet.set('Listado General de Competidores')

        file_frame = LabelFrame(self.label_frame, text='Abrir archivo')
        file_frame.place(height=100, width=400, rely=0.75, relx=0.1)

        button1 = Button(file_frame, text='Buscar un archivo', command=lambda: File_dialog())
        button1.place(rely= 0.65, relx= 0.55)

        button2 = Button(file_frame, text='Cargar archivo', command=lambda: load_excel_data())
        button2.place(rely=0.65, relx=0.30)

        label_file = ttk.Label(file_frame, text='Archivo No Seleccionado')
        label_file.place(rely=0, relx=0)

        tv1 = ttk.Treeview(frame1)
        tv1.place(relheight=1, relwidth=1)

        scrolly = Scrollbar(frame1, orient='vertical', command=tv1.yview)
        scrollx = Scrollbar(frame1, orient='horizontal', command=tv1.xview)
        tv1.configure(xscrollcommand=scrollx.set, yscrollcommand=scrolly.set)
        scrollx.pack(side='bottom', fill="x")
        scrolly.pack(side='right', fill="y")
        
        def File_dialog():
            filename = filedialog.askopenfilename( initialdir="./", title="Seleccionar Archivo", filetypes=(("xlsx files", "*.xlsx"), ("All files", "*.*") ))
            label_file["text"] = filename
            return None

        def load_excel_data():

            hojas = sheet.get()

            file_path = label_file["text"]
            try:
                excel_filename = r'{}'.format(file_path)
                if excel_filename[-4:] == ".csv":
                    df = pd.read_csv(excel_filename, sheet_name=hojas)
                else:
                    df = pd.read_excel(excel_filename, sheet_name=hojas)

            except ValueError:
                showerror("Information", "El archivo que a elegido es invalido")
                return None
            except FileNotFoundError:
                showerror("Information", f"No existe el archivo {file_path}")
                return None

            listnom(excel_filename)
            
            clear_data()
            tv1["column"] = list(df.columns)
            tv1["show"] = "headings"
            for column in tv1['column']:
                tv1.heading(column, text=column)
            
            df_rows = df.to_numpy().tolist()
            for row in df_rows:
                tv1.insert("", "end", values=row)
            return None

        def clear_data():
            tv1.delete(*tv1.get_children())
            return None

        def listnom(sheets):
            wb = load_workbook(sheets)
            wb.active
            nom_sheets = list(wb.sheetnames)
            sheet['values'] = nom_sheets
            

class CreateKeys(Frame):
    def __init__(self,master):
        Frame.__init__(self,master)
    
        # Frame buttons left
        btn_frame = Frame(self, bg='#868B8E', width=200, height=800)
        btn_frame.pack(padx=0, pady=0, side='left')
        btn_frame.grid_propagate(0)

        # Frame Labels and buttons
        label_frame = Frame(self, bg='#FFA384', width=800, height=800)
        label_frame.pack(padx=0, pady=0, side='left')
        label_frame.grid_propagate(0)   

        #Buttons Opntions
        btn_agr = Button(btn_frame, text='Agregar Competidor/a', bd=5, command=lambda: master.switch_frame(Competidor))
        btn_agr.grid(row=0, column=0, padx=25, pady=80, ipadx=5, ipady= 5)

        btn_list = Button(btn_frame, text='Listar Competidores/as', bd=5, command=lambda: master.switch_frame(ListCompet))
        btn_list.grid(row=1, column=0, padx=25, pady=(0,40), ipadx=5, ipady= 5)

        btn_word = Button(btn_frame, text='Crear Llaves', bd=5, command=lambda: master.switch_frame(CreateKeys))
        btn_word.grid(row=2, column=0, padx=25, pady=30, ipadx=5, ipady= 5)


#Cierre Pantalla
if __name__ == '__main__':
    app = App()
    app.mainloop()


